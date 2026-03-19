"""Fatture Ricevute (Received/Purchase Invoices) routes."""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel, Field
from typing import Optional, List
from enum import Enum
import uuid
import calendar
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, date, timedelta
from core.security import get_current_user
from core.database import db
from core.config import settings
from services.payment_calculator import calculate_due_dates, calc_scadenze_from_supplier
from services.audit_trail import log_activity
import logging
import re

logger = logging.getLogger(__name__)

# In-memory lock to prevent concurrent imports per user
_import_locks: set = set()
router = APIRouter(prefix="/fatture-ricevute", tags=["fatture_ricevute"])



# Ã¢ÂÂÃ¢ÂÂ Models Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

class FRStatus(str, Enum):
    DA_REGISTRARE = "da_registrare"
    REGISTRATA = "registrata"
    PAGATA = "pagata"
    CONTESTATA = "contestata"


class FRLineItem(BaseModel):
    numero_linea: int = 0
    codice_articolo: Optional[str] = None
    descrizione: str = ""
    quantita: float = 1.0
    unita_misura: str = "pz"
    prezzo_unitario: float = 0.0
    sconto_percent: float = 0.0
    aliquota_iva: str = "22"
    importo: float = 0.0


class FRCreate(BaseModel):
    fornitore_id: Optional[str] = None
    fornitore_nome: str = ""
    fornitore_piva: Optional[str] = None
    fornitore_cf: Optional[str] = None
    tipo_documento: str = "TD01"  # TD01=Fattura, TD04=Nota credito
    numero_documento: str = ""
    data_documento: str = ""  # ISO date
    data_ricezione: Optional[str] = None
    divisa: str = "EUR"
    linee: List[FRLineItem] = []
    imponibile: float = 0.0
    imposta: float = 0.0
    totale_documento: float = 0.0
    modalita_pagamento: Optional[str] = None
    condizioni_pagamento: Optional[str] = None
    data_scadenza_pagamento: Optional[str] = None
    note: Optional[str] = None
    xml_raw: Optional[str] = None
    sdi_id: Optional[str] = None


class FRUpdate(BaseModel):
    fornitore_id: Optional[str] = None
    fornitore_nome: Optional[str] = None
    status: Optional[FRStatus] = None
    note: Optional[str] = None
    linee: Optional[List[FRLineItem]] = None


class FRResponse(BaseModel):
    fr_id: str
    fornitore_id: Optional[str] = None
    fornitore_nome: str = ""
    fornitore_piva: Optional[str] = None
    fornitore_cf: Optional[str] = None
    tipo_documento: str = "TD01"
    numero_documento: str = ""
    data_documento: Optional[str] = None
    data_ricezione: Optional[str] = None
    status: str = "da_registrare"
    linee: List[dict] = []
    imponibile: float = 0.0
    imposta: float = 0.0
    totale_documento: float = 0.0
    modalita_pagamento: Optional[str] = None
    condizioni_pagamento: Optional[str] = None
    data_scadenza_pagamento: Optional[str] = None
    note: Optional[str] = None
    has_xml: bool = False
    sdi_id: Optional[str] = None
    created_at: datetime
    updated_at: Optional[datetime] = None
    # Payment tracking
    pagamenti: List[dict] = []
    totale_pagato: float = 0.0
    residuo: float = 0.0
    payment_status: str = "non_pagata"


# Ã¢ÂÂÃ¢ÂÂ FatturaPA XML Parser Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ


def _enrich_scadenze(scadenze_calc: list, origine: str) -> list:
    """Enrich payment_calculator output with full schema fields."""
    totale_rate = len(scadenze_calc)
    enriched = []
    for i, s in enumerate(scadenze_calc):
        enriched.append({
            "scadenza_id": f"scd_{uuid.uuid4().hex[:8]}",
            "numero_rata": i + 1,
            "totale_rate": totale_rate,
            "rata": s.get("rata", i + 1),
            "data_scadenza": s.get("data_scadenza", ""),
            "importo": s.get("importo", 0),
            "importo_residuo": s.get("importo", 0),
            "importo_pagato": 0.0,
            "modalita_pagamento": "",
            "stato": "aperta",
            "pagata": False,
            "origine": origine,
        })
    return enriched


NS_MAP = {
    'p': 'http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2',
}


def find_text(elem, path, default=""):
    """Find text in XML element with namespace handling."""
    if elem is None:
        return default
    # Try with namespace
    for prefix, uri in NS_MAP.items():
        full_path = path.replace('/', f'/{{{uri}}}')
        if not full_path.startswith('{'):
            full_path = f'{{{uri}}}' + full_path
        node = elem.find(full_path)
        if node is not None and node.text:
            return node.text.strip()
    # Try without namespace
    node = elem.find(path)
    if node is not None and node.text:
        return node.text.strip()
    # Try all children recursively for simple tag names
    tag_name = path.split('/')[-1]
    for child in elem.iter():
        local_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if local_name == tag_name and child.text:
            return child.text.strip()
    return default


def find_all(elem, tag_name):
    """Find all elements with given tag name (namespace-agnostic)."""
    results = []
    if elem is None:
        return results
    for child in elem.iter():
        local_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if local_name == tag_name:
            results.append(child)
    return results


def find_elem(elem, tag_name):
    """Find first element with given tag name (namespace-agnostic)."""
    if elem is None:
        return None
    for child in elem.iter():
        local_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if local_name == tag_name:
            return child
    return None



def _extract_xml_from_p7m(data: bytes) -> str | None:
    """Extract XML content from a PKCS#7 (.p7m) signed file.
    Tries multiple strategies: ASN.1 parsing, then brute-force XML extraction."""
    # Strategy 1: Try to find XML content directly in the binary data
    # FatturaPA XML always starts with <?xml and contains <FatturaElettronica
    try:
        text = data.decode('utf-8', errors='replace')
        # Find XML start
        xml_start = text.find('<?xml')
        if xml_start == -1:
            xml_start = text.find('<p:FatturaElettronica')
        if xml_start == -1:
            xml_start = text.find('<ns')
        if xml_start == -1:
            xml_start = text.find('<FatturaElettronica')
        if xml_start >= 0:
            # Find end
            for end_tag in ['</FatturaElettronica>', '</p:FatturaElettronica>',
                           '</ns2:FatturaElettronica>', '</ns3:FatturaElettronica>']:
                xml_end = text.find(end_tag, xml_start)
                if xml_end > 0:
                    return text[xml_start:xml_end + len(end_tag)]
            # Fallback: take from start to last >
            last_gt = text.rfind('>')
            if last_gt > xml_start:
                candidate = text[xml_start:last_gt + 1]
                # Validate it's parseable
                try:
                    ET.fromstring(candidate)
                    return candidate
                except ET.ParseError:
                    pass
    except Exception:
        pass

    # Strategy 2: Try OpenSSL-style extraction via subprocess
    try:
        import subprocess
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.p7m', delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        result = subprocess.run(
            ['openssl', 'smime', '-verify', '-noverify', '-in', tmp_path, '-inform', 'DER', '-out', '/dev/stdout'],
            capture_output=True, timeout=10
        )
        import os
        os.unlink(tmp_path)
        if result.stdout:
            xml_str = result.stdout.decode('utf-8', errors='replace')
            if '<FatturaElettronica' in xml_str or '<?xml' in xml_str:
                return xml_str
    except Exception:
        pass

    return None


def parse_fattura_xml(xml_content: str) -> dict:
    """Parse a FatturaPA XML and extract invoice data."""
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        raise ValueError(f"XML non valido: {str(e)}")

    # Header Ã¢ÂÂ Fornitore (CedentePrestatore)
    cedente = find_elem(root, 'CedentePrestatore')
    fornitore_piva = find_text(cedente, 'IdFiscaleIVA/IdCodice') or find_text(cedente, 'IdCodice')
    fornitore_cf = find_text(cedente, 'CodiceFiscale')
    fornitore_denominazione = find_text(cedente, 'Denominazione')
    if not fornitore_denominazione:
        nome = find_text(cedente, 'Nome')
        cognome = find_text(cedente, 'Cognome')
        fornitore_denominazione = f"{nome} {cognome}".strip()

    fornitore_sede = find_elem(cedente, 'Sede')
    fornitore_indirizzo = find_text(fornitore_sede, 'Indirizzo') if fornitore_sede else ""

    # Body Ã¢ÂÂ DatiGeneraliDocumento
    body = find_elem(root, 'FatturaElettronicaBody')
    dati_generali = find_elem(body, 'DatiGeneraliDocumento')

    tipo_doc = find_text(dati_generali, 'TipoDocumento') or "TD01"
    divisa = find_text(dati_generali, 'Divisa') or "EUR"
    data_doc = find_text(dati_generali, 'Data')
    numero_doc = find_text(dati_generali, 'Numero')
    importo_totale = find_text(dati_generali, 'ImportoTotaleDocumento')

    # Lines Ã¢ÂÂ DettaglioLinee
    linee_xml = find_all(body, 'DettaglioLinee')
    linee = []
    for i, linea in enumerate(linee_xml):
        num = find_text(linea, 'NumeroLinea') or str(i + 1)
        desc = find_text(linea, 'Descrizione')
        qty_str = find_text(linea, 'Quantita') or "1"
        um = find_text(linea, 'UnitaMisura') or "pz"
        prezzo_str = find_text(linea, 'PrezzoUnitario') or "0"
        importo_str = find_text(linea, 'PrezzoTotale') or "0"
        iva_str = find_text(linea, 'AliquotaIVA') or "22.00"

        # Parse codice articolo
        codice_art = ""
        codice_elems = find_all(linea, 'CodiceArticolo')
        if codice_elems:
            codice_art = find_text(codice_elems[0], 'CodiceValore')

        # Parse sconto
        sconto = 0.0
        sconto_elem = find_elem(linea, 'ScontoMaggiorazione')
        if sconto_elem:
            sc_perc = find_text(sconto_elem, 'Percentuale')
            if sc_perc:
                sconto = float(sc_perc)

        linee.append({
            "numero_linea": int(num) if num.isdigit() else i + 1,
            "codice_articolo": codice_art,
            "descrizione": desc,
            "quantita": float(qty_str),
            "unita_misura": um.lower(),
            "prezzo_unitario": abs(float(prezzo_str)),
            "sconto_percent": sconto,
            "aliquota_iva": iva_str.replace('.00', ''),
            "importo": float(importo_str),
        })

    # Riepilogo IVA
    riepiloghi = find_all(body, 'DatiRiepilogo')
    imponibile_tot = 0.0
    imposta_tot = 0.0
    for riep in riepiloghi:
        imp_str = find_text(riep, 'ImponibileImporto') or "0"
        iva_str = find_text(riep, 'Imposta') or "0"
        imponibile_tot += float(imp_str)
        imposta_tot += float(iva_str)

    # Pagamento Ã¢ÂÂ extract ALL DettaglioPagamento entries
    dati_pagamento = find_elem(body, 'DatiPagamento')
    modalita = ""
    condizioni = ""
    data_scadenza = ""
    scadenze_xml = []
    if dati_pagamento:
        condizioni = find_text(dati_pagamento, 'CondizioniPagamento')
        # Find all DettaglioPagamento elements
        all_dettagli = find_all(dati_pagamento, 'DettaglioPagamento')
        totale_rate = len(all_dettagli)
        for idx, det in enumerate(all_dettagli):
            det_modalita = find_text(det, 'ModalitaPagamento') or ""
            det_scadenza = find_text(det, 'DataScadenzaPagamento') or ""
            det_importo_str = find_text(det, 'ImportoPagamento') or "0"
            if not modalita:
                modalita = det_modalita
            if not data_scadenza and det_scadenza:
                data_scadenza = det_scadenza
            try:
                det_importo = round(float(det_importo_str), 2)
            except ValueError:
                det_importo = 0
            if det_scadenza:
                scadenze_xml.append({
                    "scadenza_id": f"scd_{uuid.uuid4().hex[:8]}",
                    "numero_rata": idx + 1,
                    "totale_rate": totale_rate,
                    "rata": idx + 1,
                    "data_scadenza": det_scadenza,
                    "importo": det_importo,
                    "importo_residuo": det_importo,
                    "importo_pagato": 0.0,
                    "modalita_pagamento": det_modalita,
                    "stato": "aperta",
                    "pagata": False,
                    "origine": "xml",
                })

    totale = float(importo_totale) if importo_totale else round(imponibile_tot + imposta_tot, 2)

    return {
        "fornitore_nome": fornitore_denominazione,
        "fornitore_piva": fornitore_piva,
        "fornitore_cf": fornitore_cf,
        "fornitore_indirizzo": fornitore_indirizzo,
        "tipo_documento": tipo_doc,
        "numero_documento": numero_doc,
        "data_documento": data_doc,
        "divisa": divisa,
        "linee": linee,
        "imponibile": round(imponibile_tot, 2),
        "imposta": round(imposta_tot, 2),
        "totale_documento": round(totale, 2),
        "modalita_pagamento": modalita,
        "condizioni_pagamento": condizioni,
        "data_scadenza_pagamento": data_scadenza,
        "scadenze_xml": scadenze_xml,
    }


# Ã¢ÂÂÃ¢ÂÂ CRUD Endpoints Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.get("/")
async def list_fatture_ricevute(
    q: Optional[str] = None,
    status: Optional[FRStatus] = None,
    year: Optional[int] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    user: dict = Depends(get_current_user)
):
    """List received invoices."""
    query = {"user_id": user["user_id"]}
    if status:
        query["status"] = status.value
    if year:
        query["data_documento"] = {"$regex": f"^{year}"}
    if q:
        query["$or"] = [
            {"fornitore_nome": {"$regex": q, "$options": "i"}},
            {"numero_documento": {"$regex": q, "$options": "i"}},
        ]

    total = await db.fatture_ricevute.count_documents(query)
    cursor = db.fatture_ricevute.find(query, {"_id": 0, "xml_raw": 0}).skip(skip).limit(limit).sort("created_at", -1)
    items = await cursor.to_list(length=limit)

    # Calculate KPIs
    all_query = {"user_id": user["user_id"]}
    if year:
        all_query["data_documento"] = {"$regex": f"^{year}"}
    pipeline = [
        {"$match": all_query},
        {"$group": {
            "_id": None,
            "totale": {"$sum": "$totale_documento"},
            "pagato": {"$sum": {"$ifNull": ["$totale_pagato", 0]}},
            "count": {"$sum": 1},
        }}
    ]
    agg = await db.fatture_ricevute.aggregate(pipeline).to_list(1)
    kpi = agg[0] if agg else {"totale": 0, "pagato": 0, "count": 0}

    return {
        "fatture": items,
        "total": total,
        "kpi": {
            "totale_fatture": round(kpi.get("totale", 0), 2),
            "totale_pagato": round(kpi.get("pagato", 0), 2),
            "da_pagare": round(kpi.get("totale", 0) - kpi.get("pagato", 0), 2),
            "count": kpi.get("count", 0),
        }
    }


@router.get("/{fr_id}")
async def get_fattura_ricevuta(
    fr_id: str,
    user: dict = Depends(get_current_user)
):
    """Get a single received invoice."""
    item = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]},
        {"_id": 0, "xml_raw": 0}
    )
    if not item:
        raise HTTPException(404, "Fattura ricevuta non trovata")
    return item


@router.post("/", status_code=201)
async def create_fattura_ricevuta(
    data: FRCreate,
    user: dict = Depends(get_current_user)
):
    """Create a received invoice manually."""
    now = datetime.now(timezone.utc)
    fr_id = f"fr_{uuid.uuid4().hex[:12]}"

    # Try to link supplier
    fornitore_id = data.fornitore_id
    if not fornitore_id and data.fornitore_piva:
        supplier = await db.clients.find_one(
            {"user_id": user["user_id"], "partita_iva": data.fornitore_piva},
            {"_id": 0, "client_id": 1}
        )
        if supplier:
            fornitore_id = supplier["client_id"]

    doc = {
        "fr_id": fr_id,
        "user_id": user["user_id"],
        "fornitore_id": fornitore_id,
        "fornitore_nome": data.fornitore_nome,
        "fornitore_piva": data.fornitore_piva,
        "fornitore_cf": data.fornitore_cf,
        "tipo_documento": data.tipo_documento,
        "numero_documento": data.numero_documento,
        "data_documento": data.data_documento,
        "data_ricezione": data.data_ricezione or now.strftime("%Y-%m-%d"),
        "status": FRStatus.DA_REGISTRARE.value,
        "linee": [l.model_dump() for l in data.linee],
        "imponibile": data.imponibile,
        "imposta": data.imposta,
        "totale_documento": data.totale_documento,
        "modalita_pagamento": data.modalita_pagamento,
        "condizioni_pagamento": data.condizioni_pagamento,
        "data_scadenza_pagamento": data.data_scadenza_pagamento,
        "scadenze_pagamento": [],
        "note": data.note,
        "has_xml": bool(data.xml_raw),
        "xml_raw": data.xml_raw,
        "sdi_id": data.sdi_id,
        "pagamenti": [],
        "totale_pagato": 0.0,
        "residuo": data.totale_documento,
        "payment_status": "non_pagata",
        "created_at": now,
        "updated_at": now,
    }
    # Auto-calculate scadenze from supplier payment terms (Level 2 fallback)
    if not doc["data_scadenza_pagamento"] and fornitore_id:
        scadenze_calc = await calc_scadenze_from_supplier(db, fornitore_id, user["user_id"], data.data_documento, data.totale_documento)
        if scadenze_calc:
            doc["scadenze_pagamento"] = scadenze_calc
            doc["data_scadenza_pagamento"] = scadenze_calc[-1]["data_scadenza"]
            logger.info(f"Auto-calculated {len(scadenze_calc)} scadenze for FR {fr_id} from supplier payment terms")

    await db.fatture_ricevute.insert_one(doc)
    created = await db.fatture_ricevute.find_one({"fr_id": fr_id}, {"_id": 0, "xml_raw": 0})

    # Smart Import: auto-detect welding consumables
    try:
        from routes.consumables import analyze_and_import_invoice_consumables
        consumable_doc = {**doc, "fattura_id": fr_id}
        consumables = await analyze_and_import_invoice_consumables(consumable_doc, user["user_id"])
        if consumables:
            logger.info(f"Auto-imported {len(consumables)} consumables from invoice {fr_id}")
    except Exception as e:
        logger.warning(f"Consumable auto-import failed for {fr_id}: {e}")

    return created


@router.put("/{fr_id}")
async def update_fattura_ricevuta(
    fr_id: str,
    data: FRUpdate,
    user: dict = Depends(get_current_user)
):
    """Update a received invoice."""
    existing = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]}, {"_id": 0}
    )
    if not existing:
        raise HTTPException(404, "Fattura ricevuta non trovata")

    update_dict = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if "linee" in update_dict:
        update_dict["linee"] = [l.model_dump() if hasattr(l, 'model_dump') else l for l in update_dict["linee"]]
    if "status" in update_dict:
        update_dict["status"] = update_dict["status"].value if hasattr(update_dict["status"], 'value') else update_dict["status"]

    update_dict["updated_at"] = datetime.now(timezone.utc)
    await db.fatture_ricevute.update_one({"fr_id": fr_id}, {"$set": update_dict})

    updated = await db.fatture_ricevute.find_one({"fr_id": fr_id}, {"_id": 0, "xml_raw": 0})
    return updated


@router.delete("/{fr_id}")
async def delete_fattura_ricevuta(
    fr_id: str,
    user: dict = Depends(get_current_user)
):
    """Delete a received invoice."""
    result = await db.fatture_ricevute.delete_one(
        {"fr_id": fr_id, "user_id": user["user_id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(404, "Fattura ricevuta non trovata")
    return {"message": "Fattura eliminata"}


# Ã¢ÂÂÃ¢ÂÂ XML Import Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.post("/import-xml")
async def import_xml_fattura(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Import a FatturaPA XML file (.xml or .p7m) and create a received invoice."""
    fname = file.filename.lower()
    if not (fname.endswith('.xml') or fname.endswith('.p7m')):
        raise HTTPException(400, "Il file deve essere in formato .xml o .xml.p7m")

    content = await file.read()

    # Handle .p7m (PKCS#7 signed) Ã¢ÂÂ extract XML from the wrapper
    if fname.endswith('.p7m'):
        xml_str = _extract_xml_from_p7m(content)
        if not xml_str:
            raise HTTPException(400, "Impossibile estrarre XML dal file .p7m. Prova a caricare il file .xml non firmato.")
    else:
        xml_str = content.decode('utf-8', errors='replace')

    try:
        parsed = parse_fattura_xml(xml_str)
    except ValueError as e:
        raise HTTPException(400, str(e))

    now = datetime.now(timezone.utc)
    fr_id = f"fr_{uuid.uuid4().hex[:12]}"

    # Check for duplicate Ã¢ÂÂ robust: by (numero+piva+data) OR (piva+data+totale)
    dedup_or = []
    num_doc = parsed.get("numero_documento", "")
    piva = parsed.get("fornitore_piva", "")
    data_doc = parsed.get("data_documento", "")
    totale_doc = parsed.get("totale_documento", 0)
    if num_doc and piva:
        dedup_or.append({"numero_documento": num_doc, "fornitore_piva": piva, "data_documento": data_doc})
    if piva and data_doc and totale_doc:
        dedup_or.append({"fornitore_piva": piva, "data_documento": data_doc, "totale_documento": round(totale_doc, 2)})
    existing = None
    if dedup_or:
        existing = await db.fatture_ricevute.find_one(
            {"user_id": user["user_id"], "$or": dedup_or},
            {"_id": 0, "fr_id": 1, "numero_documento": 1}
        )
    if existing:
        raise HTTPException(
            409,
            f"Fattura giÃÂ  importata: n. {parsed.get('numero_documento', '') or existing.get('numero_documento', '')} del {data_doc} da {parsed.get('fornitore_nome', '')}"
        )

    # Try to match supplier by P.IVA or Codice Fiscale
    fornitore_id = None
    if parsed.get("fornitore_piva"):
        supplier = await db.clients.find_one(
            {"user_id": user["user_id"], "partita_iva": parsed["fornitore_piva"]},
            {"_id": 0, "client_id": 1}
        )
        if supplier:
            fornitore_id = supplier["client_id"]
    if not fornitore_id and parsed.get("fornitore_cf"):
        supplier = await db.clients.find_one(
            {"user_id": user["user_id"], "fiscal_code": parsed["fornitore_cf"]},
            {"_id": 0, "client_id": 1}
        )
        if supplier:
            fornitore_id = supplier["client_id"]

    doc = {
        "fr_id": fr_id,
        "user_id": user["user_id"],
        "fornitore_id": fornitore_id,
        "fornitore_nome": parsed.get("fornitore_nome", ""),
        "fornitore_piva": parsed.get("fornitore_piva"),
        "fornitore_cf": parsed.get("fornitore_cf"),
        "tipo_documento": parsed.get("tipo_documento", "TD01"),
        "numero_documento": parsed.get("numero_documento", ""),
        "data_documento": parsed.get("data_documento", ""),
        "data_ricezione": now.strftime("%Y-%m-%d"),
        "status": FRStatus.DA_REGISTRARE.value,
        "linee": parsed.get("linee", []),
        "imponibile": parsed.get("imponibile", 0),
        "imposta": parsed.get("imposta", 0),
        "totale_documento": parsed.get("totale_documento", 0),
        "modalita_pagamento": parsed.get("modalita_pagamento"),
        "condizioni_pagamento": parsed.get("condizioni_pagamento"),
        "data_scadenza_pagamento": parsed.get("data_scadenza_pagamento"),
        "scadenze_pagamento": [],
        "note": None,
        "has_xml": True,
        "xml_raw": xml_str,
        "sdi_id": None,
        "pagamenti": [],
        "totale_pagato": 0.0,
        "residuo": parsed.get("totale_documento", 0),
        "payment_status": "non_pagata",
        "created_at": now,
        "updated_at": now,
    }

    # Level 1: Use XML payment schedule if available
    xml_scadenze = parsed.get("scadenze_xml", [])
    scadenze_origine = "nessuna"
    if xml_scadenze:
        doc["scadenze_pagamento"] = xml_scadenze
        doc["data_scadenza_pagamento"] = xml_scadenze[-1]["data_scadenza"]
        scadenze_origine = "xml"
        logger.info(f"Level 1: {len(xml_scadenze)} scadenze from XML for FR {fr_id}")
    # Level 2: Fallback - calculate from supplier payment terms
    elif fornitore_id:
        scadenze_calc = await calc_scadenze_from_supplier(
            db, fornitore_id, user["user_id"],
            parsed.get("data_documento", ""),
            parsed.get("totale_documento", 0)
        )
        if scadenze_calc:
            doc["scadenze_pagamento"] = _enrich_scadenze(scadenze_calc, "fornitore")
            doc["data_scadenza_pagamento"] = scadenze_calc[-1]["data_scadenza"]
            scadenze_origine = "fornitore"
            logger.info(f"Level 2: {len(scadenze_calc)} scadenze from supplier for FR {fr_id}")
    # Level 3: Default 30 days from invoice date
    if not doc["scadenze_pagamento"]:
        data_doc = parsed.get("data_documento", "")
        totale = parsed.get("totale_documento", 0)
        if data_doc and totale > 0:
            try:
                d0 = date.fromisoformat(data_doc)
                default_scad = (d0 + timedelta(days=30)).isoformat()
            except (ValueError, TypeError):
                default_scad = ""
            if default_scad:
                doc["scadenze_pagamento"] = [{
                    "scadenza_id": f"scd_{uuid.uuid4().hex[:8]}",
                    "numero_rata": 1,
                    "totale_rate": 1,
                    "rata": 1,
                    "data_scadenza": default_scad,
                    "importo": round(totale, 2),
                    "importo_residuo": round(totale, 2),
                    "importo_pagato": 0.0,
                    "modalita_pagamento": "",
                    "stato": "aperta",
                    "pagata": False,
                    "origine": "default_30gg",
                }]
                doc["data_scadenza_pagamento"] = default_scad
                scadenze_origine = "default_30gg"
                logger.info(f"Level 3: default 30gg scadenza for FR {fr_id}")

    await db.fatture_ricevute.insert_one(doc)
    created = await db.fatture_ricevute.find_one({"fr_id": fr_id}, {"_id": 0, "xml_raw": 0})

    # Smart Import: auto-detect welding consumables
    try:
        from routes.consumables import analyze_and_import_invoice_consumables
        consumable_doc = {**doc, "fattura_id": fr_id}
        consumables = await analyze_and_import_invoice_consumables(consumable_doc, user["user_id"])
        if consumables:
            logger.info(f"Auto-imported {len(consumables)} consumables from XML invoice {fr_id}")
    except Exception as e:
        logger.warning(f"Consumable auto-import failed for XML {fr_id}: {e}")

    return {
        "message": f"Fattura importata: {parsed.get('numero_documento', 'N/A')} da {parsed.get('fornitore_nome', 'N/A')} Ã¢ÂÂ {parsed.get('totale_documento', 0):.2f}Ã¢ÂÂ¬",
        "fattura": created,
        "fornitore_trovato": fornitore_id is not None,
        "scadenze_origine": scadenze_origine,
        "scadenze_count": len(doc.get("scadenze_pagamento", [])),
    }


@router.post("/import-xml-batch")
async def import_xml_batch(
    files: List[UploadFile] = File(...),
    user: dict = Depends(get_current_user)
):
    """Import multiple FatturaPA XML files at once."""
    await log_activity(user, "import", "fattura_ricevuta", "", label=f"Batch import {len(files)} file XML")
    results = {"imported": 0, "skipped": 0, "errors": [], "fatture": [], "dettaglio_saltate": []}

    for f in files:
        fname = f.filename.lower()
        if not (fname.endswith('.xml') or fname.endswith('.p7m')):
            results["errors"].append(f"{f.filename}: formato non supportato (solo .xml o .p7m)")
            continue

        content = await f.read()
        if fname.endswith('.p7m'):
            xml_str = _extract_xml_from_p7m(content)
            if not xml_str:
                results["errors"].append(f"{f.filename}: impossibile estrarre XML dal .p7m")
                continue
        else:
            xml_str = content.decode('utf-8', errors='replace')

        try:
            parsed = parse_fattura_xml(xml_str)
        except ValueError as e:
            results["errors"].append(f"{f.filename}: {str(e)}")
            continue

        # Check duplicate Ã¢ÂÂ robust: by (numero+piva+data) OR (piva+data+totale)
        dedup_or = []
        num_doc = parsed.get("numero_documento", "")
        b_piva = parsed.get("fornitore_piva", "")
        b_data = parsed.get("data_documento", "")
        b_totale = parsed.get("totale_documento", 0)
        if num_doc and b_piva:
            dedup_or.append({"numero_documento": num_doc, "fornitore_piva": b_piva, "data_documento": b_data})
        if b_piva and b_data and b_totale:
            dedup_or.append({"fornitore_piva": b_piva, "data_documento": b_data, "totale_documento": round(b_totale, 2)})
        existing = None
        if dedup_or:
            existing = await db.fatture_ricevute.find_one(
                {"user_id": user["user_id"], "$or": dedup_or}, {"_id": 0}
            )
        if existing:
            results["skipped"] += 1
            results["dettaglio_saltate"].append({
                "numero": num_doc or "N/A",
                "fornitore": parsed.get("fornitore_nome", ""),
                "data": b_data,
                "motivo": "giÃÂ  presente",
            })
            results["errors"].append(f"{f.filename}: giÃÂ  importata (n. {num_doc or 'N/A'} del {b_data})")
            continue

        now = datetime.now(timezone.utc)
        fr_id = f"fr_{uuid.uuid4().hex[:12]}"

        fornitore_id = None
        if parsed.get("fornitore_piva"):
            supplier = await db.clients.find_one(
                {"user_id": user["user_id"], "partita_iva": parsed["fornitore_piva"]},
                {"_id": 0, "client_id": 1}
            )
            if supplier:
                fornitore_id = supplier["client_id"]
        if not fornitore_id and parsed.get("fornitore_cf"):
            supplier = await db.clients.find_one(
                {"user_id": user["user_id"], "fiscal_code": parsed["fornitore_cf"]},
                {"_id": 0, "client_id": 1}
            )
            if supplier:
                fornitore_id = supplier["client_id"]

        doc = {
            "fr_id": fr_id,
            "user_id": user["user_id"],
            "fornitore_id": fornitore_id,
            "fornitore_nome": parsed.get("fornitore_nome", ""),
            "fornitore_piva": parsed.get("fornitore_piva"),
            "fornitore_cf": parsed.get("fornitore_cf"),
            "tipo_documento": parsed.get("tipo_documento", "TD01"),
            "numero_documento": parsed.get("numero_documento", ""),
            "data_documento": parsed.get("data_documento", ""),
            "data_ricezione": now.strftime("%Y-%m-%d"),
            "status": FRStatus.DA_REGISTRARE.value,
            "linee": parsed.get("linee", []),
            "imponibile": parsed.get("imponibile", 0),
            "imposta": parsed.get("imposta", 0),
            "totale_documento": parsed.get("totale_documento", 0),
            "modalita_pagamento": parsed.get("modalita_pagamento"),
            "condizioni_pagamento": parsed.get("condizioni_pagamento"),
            "data_scadenza_pagamento": parsed.get("data_scadenza_pagamento"),
            "scadenze_pagamento": [],
            "note": None,
            "has_xml": True,
            "xml_raw": xml_str,
            "sdi_id": None,
            "pagamenti": [],
            "totale_pagato": 0.0,
            "residuo": parsed.get("totale_documento", 0),
            "payment_status": "non_pagata",
            "created_at": now,
            "updated_at": now,
        }

        # Level 1: XML scadenze
        xml_scadenze = parsed.get("scadenze_xml", [])
        if xml_scadenze:
            doc["scadenze_pagamento"] = xml_scadenze
            doc["data_scadenza_pagamento"] = xml_scadenze[-1]["data_scadenza"]
        # Level 2: Supplier fallback
        elif fornitore_id:
            scadenze_calc = await calc_scadenze_from_supplier(
                db, fornitore_id, user["user_id"],
                parsed.get("data_documento", ""),
                parsed.get("totale_documento", 0)
            )
            if scadenze_calc:
                doc["scadenze_pagamento"] = _enrich_scadenze(scadenze_calc, "fornitore")
                doc["data_scadenza_pagamento"] = scadenze_calc[-1]["data_scadenza"]
        # Level 3: Default 30 days
        if not doc["scadenze_pagamento"]:
            data_doc = parsed.get("data_documento", "")
            totale = parsed.get("totale_documento", 0)
            if data_doc and totale > 0:
                try:
                    d0 = date.fromisoformat(data_doc)
                    default_scad = (d0 + timedelta(days=30)).isoformat()
                except (ValueError, TypeError):
                    default_scad = ""
                if default_scad:
                    doc["scadenze_pagamento"] = [{
                        "scadenza_id": f"scd_{uuid.uuid4().hex[:8]}",
                        "numero_rata": 1, "totale_rate": 1, "rata": 1,
                        "data_scadenza": default_scad,
                        "importo": round(totale, 2), "importo_residuo": round(totale, 2),
                        "importo_pagato": 0.0, "modalita_pagamento": "",
                        "stato": "aperta", "pagata": False, "origine": "default_30gg",
                    }]
                    doc["data_scadenza_pagamento"] = default_scad

        await db.fatture_ricevute.insert_one(doc)
        results["imported"] += 1
        results["fatture"].append({
            "filename": f.filename,
            "numero": parsed.get("numero_documento", ""),
            "fornitore": parsed.get("fornitore_nome", ""),
            "totale": parsed.get("totale_documento", 0),
        })

    return {
        "message": f"Importazione completata: {results['imported']} importate, {results['skipped']} duplicate",
        **results,
    }


# Ã¢ÂÂÃ¢ÂÂ Parse XML Preview (no save) Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.post("/preview-xml")
async def preview_xml_fattura(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Parse a FatturaPA XML file and return preview with payment schedule."""
    fname = file.filename.lower()
    content = await file.read()

    if fname.endswith('.p7m'):
        xml_str = _extract_xml_from_p7m(content)
        if not xml_str:
            raise HTTPException(400, "Impossibile estrarre XML dal file .p7m")
    else:
        xml_str = content.decode('utf-8', errors='replace')

    try:
        parsed = parse_fattura_xml(xml_str)
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Calculate payment schedule with fallback
    scadenze_preview = parsed.get("scadenze_xml", [])
    scadenze_origine = "xml" if scadenze_preview else "nessuna"

    if not scadenze_preview:
        # Try supplier fallback
        fornitore_id = None
        if parsed.get("fornitore_piva"):
            supplier = await db.clients.find_one(
                {"user_id": user["user_id"], "partita_iva": parsed["fornitore_piva"]},
                {"_id": 0, "client_id": 1, "business_name": 1}
            )
            if supplier:
                fornitore_id = supplier["client_id"]
        if not fornitore_id and parsed.get("fornitore_cf"):
            supplier = await db.clients.find_one(
                {"user_id": user["user_id"], "fiscal_code": parsed["fornitore_cf"]},
                {"_id": 0, "client_id": 1, "business_name": 1}
            )
            if supplier:
                fornitore_id = supplier["client_id"]

        if fornitore_id:
            scadenze_calc = await calc_scadenze_from_supplier(
                db, fornitore_id, user["user_id"],
                parsed.get("data_documento", ""),
                parsed.get("totale_documento", 0)
            )
            if scadenze_calc:
                scadenze_preview = _enrich_scadenze(scadenze_calc, "fornitore")
                scadenze_origine = "fornitore"

        if not scadenze_preview:
            data_doc = parsed.get("data_documento", "")
            totale = parsed.get("totale_documento", 0)
            if data_doc and totale > 0:
                try:
                    d0 = date.fromisoformat(data_doc)
                    default_scad = (d0 + timedelta(days=30)).isoformat()
                    scadenze_preview = [{
                        "scadenza_id": "preview",
                        "numero_rata": 1, "totale_rate": 1, "rata": 1,
                        "data_scadenza": default_scad,
                        "importo": round(totale, 2), "importo_residuo": round(totale, 2),
                        "importo_pagato": 0.0, "modalita_pagamento": "",
                        "stato": "aperta", "pagata": False, "origine": "default_30gg",
                    }]
                    scadenze_origine = "default_30gg"
                except (ValueError, TypeError):
                    pass

    # Check for existing duplicate Ã¢ÂÂ robust
    existing = None
    dedup_or = []
    p_num = parsed.get("numero_documento", "")
    p_piva = parsed.get("fornitore_piva", "")
    p_data = parsed.get("data_documento", "")
    p_totale = parsed.get("totale_documento", 0)
    if p_num and p_piva:
        dedup_or.append({"numero_documento": p_num, "fornitore_piva": p_piva, "data_documento": p_data})
    if p_piva and p_data and p_totale:
        dedup_or.append({"fornitore_piva": p_piva, "data_documento": p_data, "totale_documento": round(p_totale, 2)})
    if dedup_or:
        existing = await db.fatture_ricevute.find_one(
            {"user_id": user["user_id"], "$or": dedup_or},
            {"_id": 0, "fr_id": 1}
        )

    return {
        "preview": parsed,
        "scadenze_calcolate": scadenze_preview,
        "scadenze_origine": scadenze_origine,
        "fornitore_trovato": fornitore_id is not None if not scadenze_preview else None,
        "duplicata": existing is not None,
    }


# Ã¢ÂÂÃ¢ÂÂ Extract Articles to Catalog Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.post("/{fr_id}/extract-articoli")
async def extract_articoli(
    fr_id: str,
    user: dict = Depends(get_current_user)
):
    """Extract line items from a received invoice into the Catalogo Articoli."""
    fr = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]},
        {"_id": 0}
    )
    if not fr:
        raise HTTPException(404, "Fattura ricevuta non trovata")

    linee = fr.get("linee", [])
    fornitore_nome = fr.get("fornitore_nome", "")
    now = datetime.now(timezone.utc)

    created_count = 0
    updated_count = 0
    skipped_count = 0

    for linea in linee:
        desc = linea.get("descrizione", "").strip()
        if not desc or len(desc) < 3:
            skipped_count += 1
            continue

        prezzo = abs(linea.get("prezzo_unitario", 0))
        codice = (linea.get("codice_articolo") or "").strip()
        if not codice:
            # Generate code from first 3 words of description
            words = re.sub(r'[^a-zA-Z0-9\s]', '', desc).upper().split()[:3]
            codice = "-".join(words) if words else f"ART-{uuid.uuid4().hex[:4].upper()}"

        um = (linea.get("unita_misura") or "pz").lower()
        if um not in ["pz", "ml", "mq", "kg", "h", "corpo", "lt"]:
            um = "pz"

        iva = (linea.get("aliquota_iva") or "22").replace('.00', '')

        # Check existing
        existing = await db.articoli.find_one(
            {"user_id": user["user_id"], "codice": codice},
            {"_id": 0}
        )

        if existing:
            # Update price if different
            if abs(prezzo - existing.get("prezzo_unitario", 0)) > 0.01:
                await db.articoli.update_one(
                    {"articolo_id": existing["articolo_id"]},
                    {
                        "$set": {"prezzo_unitario": prezzo, "updated_at": now},
                        "$push": {"storico_prezzi": {
                            "prezzo": prezzo,
                            "data": now.isoformat(),
                            "fonte": f"Fatt. {fr.get('numero_documento', '')} Ã¢ÂÂ {fornitore_nome}"
                        }}
                    }
                )
                updated_count += 1
            else:
                skipped_count += 1
        else:
            doc = {
                "articolo_id": f"art_{uuid.uuid4().hex[:12]}",
                "user_id": user["user_id"],
                "codice": codice,
                "descrizione": desc,
                "categoria": "materiale",
                "unita_misura": um,
                "prezzo_unitario": prezzo,
                "aliquota_iva": iva,
                "fornitore_nome": fornitore_nome,
                "fornitore_id": fr.get("fornitore_id"),
                "note": f"Importato da fattura {fr.get('numero_documento', '')}",
                "storico_prezzi": [{"prezzo": prezzo, "data": now.isoformat(), "fonte": f"Fatt. {fr.get('numero_documento', '')} Ã¢ÂÂ {fornitore_nome}"}],
                "created_at": now,
                "updated_at": now,
            }
            await db.articoli.insert_one(doc)
            created_count += 1

    return {
        "message": f"Estrazione completata: {created_count} creati, {updated_count} aggiornati, {skipped_count} saltati",
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
    }


# Ã¢ÂÂÃ¢ÂÂ Payment Tracking Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

class FRPayment(BaseModel):
    importo: float = Field(..., gt=0)
    data_pagamento: str
    metodo: Optional[str] = None
    note: Optional[str] = None


@router.get("/{fr_id}/pagamenti")
async def get_fr_pagamenti(
    fr_id: str,
    user: dict = Depends(get_current_user)
):
    """Get payment info for a received invoice."""
    fr = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]},
        {"_id": 0, "xml_raw": 0}
    )
    if not fr:
        raise HTTPException(404, "Fattura ricevuta non trovata")

    pagamenti = fr.get("pagamenti", [])
    pagato = sum(p.get("importo", 0) for p in pagamenti)
    total_doc = fr.get("totale_documento", 0)
    residuo = round(total_doc - pagato, 2)

    return {
        "fr_id": fr_id,
        "total_document": total_doc,
        "pagamenti": pagamenti,
        "totale_pagato": round(pagato, 2),
        "residuo": max(residuo, 0),
        "payment_status": "pagata" if residuo <= 0.01 else ("parzialmente_pagata" if pagato > 0 else "non_pagata"),
    }


@router.post("/{fr_id}/pagamenti")
async def record_fr_payment(
    fr_id: str,
    payment: FRPayment,
    user: dict = Depends(get_current_user)
):
    """Record payment for a received invoice."""
    fr = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]},
        {"_id": 0}
    )
    if not fr:
        raise HTTPException(404, "Fattura ricevuta non trovata")

    pagamenti = fr.get("pagamenti", [])
    paid_so_far = sum(p.get("importo", 0) for p in pagamenti)
    total_doc = fr.get("totale_documento", 0)

    if payment.importo > (total_doc - paid_so_far + 0.01):
        raise HTTPException(400, "Importo supera il residuo")

    new_payment = {
        "payment_id": f"pay_{uuid.uuid4().hex[:8]}",
        "importo": payment.importo,
        "data_pagamento": payment.data_pagamento,
        "metodo": payment.metodo,
        "note": payment.note,
        "recorded_at": datetime.now(timezone.utc).isoformat(),
    }

    new_paid = paid_so_far + payment.importo
    residuo = round(total_doc - new_paid, 2)
    ps = "pagata" if residuo <= 0.01 else "parzialmente_pagata"

    update = {
        "$push": {"pagamenti": new_payment},
        "$set": {
            "totale_pagato": round(new_paid, 2),
            "residuo": max(residuo, 0),
            "payment_status": ps,
            "updated_at": datetime.now(timezone.utc),
        }
    }
    if residuo <= 0.01:
        update["$set"]["status"] = "pagata"

    await db.fatture_ricevute.update_one({"fr_id": fr_id}, update)
    return {"message": "Pagamento registrato", "totale_pagato": round(new_paid, 2), "residuo": max(residuo, 0), "payment_status": ps}



# Ã¢ÂÂÃ¢ÂÂ Cost Imputation (Assign to Commessa or Magazzino) Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

class ImputazioneDestinazione(str, Enum):
    COMMESSA = "commessa"
    MAGAZZINO = "magazzino"


class ImputazioneRequest(BaseModel):
    destinazione: ImputazioneDestinazione
    commessa_id: Optional[str] = None
    righe_selezionate: Optional[List[int]] = None  # line indexes, None = all
    note: Optional[str] = None


@router.post("/{fr_id}/imputa")
async def imputa_costi(
    fr_id: str,
    data: ImputazioneRequest,
    user: dict = Depends(get_current_user)
):
    """Assign invoice costs to a commessa or magazzino."""
    fr = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]}, {"_id": 0, "xml_raw": 0}
    )
    if not fr:
        raise HTTPException(404, "Fattura ricevuta non trovata")

    linee = fr.get("linee", [])
    if data.righe_selezionate:
        selected = [linee[i] for i in data.righe_selezionate if i < len(linee)]
    else:
        selected = linee

    if not selected:
        raise HTTPException(400, "Nessuna riga selezionata")

    totale_imputato = sum(abs(l.get("importo", 0)) for l in selected)
    now = datetime.now(timezone.utc)

    if data.destinazione == ImputazioneDestinazione.COMMESSA:
        if not data.commessa_id:
            raise HTTPException(400, "commessa_id obbligatorio per destinazione commessa")

        commessa = await db.commesse.find_one(
            {"commessa_id": data.commessa_id, "user_id": user["user_id"]},
            {"_id": 0, "commessa_id": 1, "numero": 1}
        )
        if not commessa:
            raise HTTPException(404, "Commessa non trovata")

        # Add cost entry to commessa
        cost_entry = {
            "cost_id": f"cost_{uuid.uuid4().hex[:8]}",
            "tipo": "materiale",
            "descrizione": f"Fatt. {fr.get('numero_documento', '')} Ã¢ÂÂ {fr.get('fornitore_nome', '')}",
            "fornitore": fr.get("fornitore_nome", ""),
            "importo": round(totale_imputato, 2),
            "data": now.isoformat(),
            "fr_id": fr_id,
            "note": data.note or "",
            "righe": [{
                "descrizione": l.get("descrizione", ""),
                "quantita": l.get("quantita", 0),
                "prezzo_unitario": l.get("prezzo_unitario", 0),
                "importo": l.get("importo", 0),
            } for l in selected],
        }

        await db.commesse.update_one(
            {"commessa_id": data.commessa_id},
            {"$push": {"costi_reali": cost_entry}, "$set": {"updated_at": now}}
        )

        # Mark fattura as processed
        await db.fatture_ricevute.update_one(
            {"fr_id": fr_id},
            {"$set": {
                "imputazione": {
                    "destinazione": "commessa",
                    "commessa_id": data.commessa_id,
                    "commessa_numero": commessa.get("numero", ""),
                    "importo": round(totale_imputato, 2),
                    "data": now.isoformat(),
                },
                "status": "registrata",
                "updated_at": now,
            }}
        )

        return {
            "message": f"Costo di {totale_imputato:.2f}Ã¢ÂÂ¬ imputato alla commessa {commessa.get('numero', '')}",
            "destinazione": "commessa",
            "commessa_numero": commessa.get("numero", ""),
            "importo": round(totale_imputato, 2),
        }

    else:  # MAGAZZINO
        updated_count = 0
        created_count = 0
        for linea in selected:
            desc = linea.get("descrizione", "").strip()
            if not desc or len(desc) < 3:
                continue

            prezzo = abs(linea.get("prezzo_unitario", 0))
            qty = abs(linea.get("quantita", 0))
            codice = (linea.get("codice_articolo") or "").strip()

            if codice:
                existing = await db.articoli.find_one(
                    {"user_id": user["user_id"], "codice": codice}, {"_id": 0}
                )
            else:
                existing = None

            if existing:
                # Update stock and price
                old_price = existing.get("prezzo_unitario", 0)
                old_stock = existing.get("giacenza", 0)
                new_stock = old_stock + qty

                # Weighted average price
                if old_stock + qty > 0:
                    new_price = ((old_price * old_stock) + (prezzo * qty)) / (old_stock + qty)
                else:
                    new_price = prezzo

                await db.articoli.update_one(
                    {"articolo_id": existing["articolo_id"]},
                    {
                        "$set": {
                            "prezzo_unitario": round(new_price, 4),
                            "giacenza": new_stock,
                            "updated_at": now,
                        },
                        "$push": {"storico_prezzi": {
                            "prezzo": prezzo,
                            "data": now.isoformat(),
                            "fonte": f"Fatt. {fr.get('numero_documento', '')} Ã¢ÂÂ {fr.get('fornitore_nome', '')}",
                            "quantita": qty,
                        }}
                    }
                )
                updated_count += 1
            else:
                # Create new article with stock
                words = re.sub(r'[^a-zA-Z0-9\s]', '', desc).upper().split()[:3]
                auto_codice = "-".join(words) if words else f"ART-{uuid.uuid4().hex[:4].upper()}"
                um = (linea.get("unita_misura") or "pz").lower()

                doc = {
                    "articolo_id": f"art_{uuid.uuid4().hex[:12]}",
                    "user_id": user["user_id"],
                    "codice": codice or auto_codice,
                    "descrizione": desc,
                    "categoria": "materiale",
                    "unita_misura": um,
                    "prezzo_unitario": prezzo,
                    "giacenza": qty,
                    "aliquota_iva": (linea.get("aliquota_iva") or "22").replace('.00', ''),
                    "fornitore_nome": fr.get("fornitore_nome", ""),
                    "fornitore_id": fr.get("fornitore_id"),
                    "storico_prezzi": [{"prezzo": prezzo, "data": now.isoformat(),
                        "fonte": f"Fatt. {fr.get('numero_documento', '')} Ã¢ÂÂ {fr.get('fornitore_nome', '')}",
                        "quantita": qty}],
                    "created_at": now,
                    "updated_at": now,
                }
                await db.articoli.insert_one(doc)
                created_count += 1

        # Mark fattura as processed
        await db.fatture_ricevute.update_one(
            {"fr_id": fr_id},
            {"$set": {
                "imputazione": {
                    "destinazione": "magazzino",
                    "importo": round(totale_imputato, 2),
                    "data": now.isoformat(),
                    "articoli_aggiornati": updated_count,
                    "articoli_creati": created_count,
                },
                "status": "registrata",
                "updated_at": now,
            }}
        )

        return {
            "message": f"Magazzino aggiornato: {created_count} articoli creati, {updated_count} aggiornati",
            "destinazione": "magazzino",
            "importo": round(totale_imputato, 2),
            "created": created_count,
            "updated": updated_count,
        }



# Ã¢ÂÂÃ¢ÂÂ Annulla Imputazione (Undo cost assignment) Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.post("/{fr_id}/annulla-imputazione")
async def annulla_imputazione(
    fr_id: str,
    user: dict = Depends(get_current_user)
):
    """Undo the cost assignment of a received invoice from a commessa.
    Removes the cost entry from the commessa and clears the imputazione on the invoice."""
    fr = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]},
        {"_id": 0, "xml_raw": 0}
    )
    if not fr:
        raise HTTPException(404, "Fattura ricevuta non trovata")
    
    imputazione = fr.get("imputazione")
    if not imputazione:
        raise HTTPException(400, "Questa fattura non ha nessuna imputazione da annullare")
    
    destinazione = imputazione.get("destinazione", "")
    commessa_id = imputazione.get("commessa_id", "")
    
    now = datetime.now(timezone.utc)
    
    # If assigned to a commessa, remove the cost entry
    if destinazione == "commessa" and commessa_id:
        # Remove cost entries from commessa that reference this fr_id
        result = await db.commesse.update_one(
            {"commessa_id": commessa_id, "user_id": user["user_id"]},
            {
                "$pull": {"costi_reali": {"fr_id": fr_id}},
                "$set": {"updated_at": now},
            }
        )
        commessa = await db.commesse.find_one(
            {"commessa_id": commessa_id, "user_id": user["user_id"]},
            {"_id": 0, "numero": 1}
        )
        commessa_numero = commessa.get("numero", commessa_id) if commessa else commessa_id
        
        logger.info(f"Annullata imputazione fattura {fr_id} dalla commessa {commessa_id}")
    elif destinazione == "magazzino":
        # If assigned to magazzino, we could reverse the stock changes
        # but for simplicity, we just remove the imputazione flag
        commessa_numero = "Magazzino"
        logger.info(f"Annullata imputazione fattura {fr_id} dal magazzino")
    else:
        commessa_numero = destinazione
    
    # Clear imputazione on the invoice, revert status to da_registrare
    await db.fatture_ricevute.update_one(
        {"fr_id": fr_id},
        {
            "$unset": {"imputazione": ""},
            "$set": {
                "status": "da_registrare",
                "updated_at": now,
            }
        }
    )
    
    return {
        "message": f"Imputazione annullata. La fattura non ÃÂ¨ piÃÂ¹ collegata a {commessa_numero}.",
        "fr_id": fr_id,
        "previous_destinazione": destinazione,
        "previous_commessa_id": commessa_id,
    }


# Ã¢ÂÂÃ¢ÂÂ FattureInCloud Sync Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.post("/sync-fic")
async def sync_fatture_from_fic(
    user: dict = Depends(get_current_user)
):
    """Sync received invoices from FattureInCloud API."""
    uid = user["user_id"]
    if uid in _import_locks:
        raise HTTPException(429, "Import giÃÂ  in corso, attendi il completamento")
    _import_locks.add(uid)
    try:
        return await _sync_fatture_from_fic_impl(user)
    finally:
        _import_locks.discard(uid)


async def _sync_fatture_from_fic_impl(user: dict):
    from services.fattureincloud_api import get_fic_client

    # Get user's FIC credentials from settings or user profile
    user_doc = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})
    fic_token = (user_doc or {}).get("fic_access_token") or getattr(settings, 'fic_access_token', None)
    fic_company_id = (user_doc or {}).get("fic_company_id") or getattr(settings, 'fic_company_id', None)

    if not fic_token or not fic_company_id:
        raise HTTPException(400, "FattureInCloud non configurato. Inserisci il token API nelle impostazioni.")

    client = get_fic_client(access_token=fic_token, company_id=int(fic_company_id))
    if not client.is_configured:
        raise HTTPException(400, "Configurazione FattureInCloud incompleta")

    imported = 0
    skipped = 0
    errors = []
    page = 1

    # PUNTO 1 Ã¢ÂÂ Leggi watermark ultima sync
    sync_state = await db.sync_state.find_one(
        {"user_id": user["user_id"], "type": "fatture_ricevute"},
        {"_id": 0, "last_sync_at": 1}
    )
    last_sync_date = None
    if sync_state and sync_state.get("last_sync_at"):
        last_sync_date = sync_state["last_sync_at"][:10]

    try:
        while True:
            # PUNTO 2 Ã¢ÂÂ Filtra per data se watermark disponibile
            extra_params = {}
            if last_sync_date:
                extra_params["filter[date][from]"] = last_sync_date
            resp = await client.list_received_invoices(page=page, per_page=50, **extra_params)
            data_list = resp.get("data", [])
            if not data_list:
                break

            for doc_fic in data_list:
                fic_id = str(doc_fic.get("id", ""))
                # Skip if already imported (by fic_id OR by fingerprint)
                existing = await db.fatture_ricevute.find_one(
                    {"user_id": user["user_id"], "fic_id": fic_id}, {"_id": 0, "fr_id": 1}
                )
                if not existing:
                    # Also check by fingerprint: piva + date + total
                    doc_date = str(doc_fic.get("date", ""))
                    doc_total = round(float(doc_fic.get("amount_gross", 0)), 2)
                    entity_piva = (doc_fic.get("entity") or {}).get("vat_number", "")
                    if entity_piva and doc_date:
                        existing = await db.fatture_ricevute.find_one({
                            "user_id": user["user_id"],
                            "fornitore_piva": entity_piva,
                            "data_documento": doc_date,
                            "totale_documento": doc_total,
                        }, {"_id": 0, "fr_id": 1})
                if existing:
                    skipped += 1
                    continue

                # Fetch detailed document to get items_list
                try:
                    detail_resp = await client.get_received_document_detail(fic_id)
                    doc_detail = detail_resp.get("data", {})
                except Exception:
                    doc_detail = {}

                # Map FIC document to our schema
                entity = doc_fic.get("entity", {}) or {}
                items = doc_detail.get("items_list") or doc_fic.get("items_list") or []

                linee = []
                for i, item in enumerate(items):
                    linee.append({
                        "numero_linea": i + 1,
                        "codice_articolo": item.get("code", ""),
                        "descrizione": item.get("name", ""),
                        "quantita": float(item.get("qty", 1)),
                        "unita_misura": item.get("measure", "pz") or "pz",
                        "prezzo_unitario": abs(float(item.get("net_price", 0))),
                        "sconto_percent": float(item.get("discount", 0)),
                        "aliquota_iva": str(int(item.get("vat", {}).get("value", 22))) if isinstance(item.get("vat"), dict) else "22",
                        "importo": float(item.get("net_price", 0)) * float(item.get("qty", 1)),
                    })

                amount_net = float(doc_fic.get("amount_net", 0))
                amount_vat = float(doc_fic.get("amount_vat", 0))
                amount_gross = float(doc_fic.get("amount_gross", amount_net + amount_vat))

                # Payment info Ã¢ÂÂ extract all FIC payment entries
                payments_list = doc_fic.get("payments_list", []) or []
                data_scadenza = ""
                fic_scadenze = []
                for pidx, pay in enumerate(payments_list):
                    pay_date = pay.get("due_date", "")
                    pay_amount = pay.get("amount", 0)
                    pay_paid = pay.get("status", "") == "paid"
                    if pay_date:
                        if not data_scadenza:
                            data_scadenza = pay_date
                        fic_scadenze.append({
                            "scadenza_id": f"scd_{uuid.uuid4().hex[:8]}",
                            "numero_rata": pidx + 1,
                            "totale_rate": len(payments_list),
                            "rata": pidx + 1,
                            "data_scadenza": pay_date,
                            "importo": round(pay_amount, 2) if pay_amount else 0,
                            "importo_residuo": 0 if pay_paid else round(pay_amount, 2),
                            "importo_pagato": round(pay_amount, 2) if pay_paid else 0,
                            "modalita_pagamento": "",
                            "stato": "pagata" if pay_paid else "aperta",
                            "pagata": pay_paid,
                            "origine": "fic",
                        })
                if not data_scadenza and payments_list:
                    data_scadenza = payments_list[0].get("due_date", "")

                now = datetime.now(timezone.utc)
                fr_id = f"fr_{uuid.uuid4().hex[:12]}"

                # Match supplier by VAT
                fornitore_id = None
                piva = entity.get("vat_number", "")
                if piva:
                    supplier = await db.clients.find_one(
                        {"user_id": user["user_id"], "partita_iva": piva},
                        {"_id": 0, "client_id": 1}
                    )
                    if supplier:
                        fornitore_id = supplier["client_id"]

                fr_doc = {
                    "fr_id": fr_id,
                    "fic_id": fic_id,
                    "user_id": user["user_id"],
                    "fornitore_id": fornitore_id,
                    "fornitore_nome": entity.get("name", ""),
                    "fornitore_piva": piva,
                    "fornitore_cf": entity.get("tax_code", ""),
                    "tipo_documento": "TD01",
                    "numero_documento": str(doc_fic.get("number", "")),
                    "data_documento": str(doc_fic.get("date", "")),
                    "data_ricezione": now.strftime("%Y-%m-%d"),
                    "status": "pagata" if doc_fic.get("is_marked") else "da_registrare",
                    "linee": linee,
                    "imponibile": round(amount_net, 2),
                    "imposta": round(amount_vat, 2),
                    "totale_documento": round(amount_gross, 2),
                    "modalita_pagamento": "",
                    "condizioni_pagamento": "",
                    "data_scadenza_pagamento": data_scadenza,
                    "scadenze_pagamento": fic_scadenze,
                    "note": doc_fic.get("notes", ""),
                    "has_xml": False,
                    "sdi_id": None,
                    "pagamenti": [],
                    "totale_pagato": round(amount_gross, 2) if doc_fic.get("is_marked") else 0.0,
                    "residuo": 0.0 if doc_fic.get("is_marked") else round(amount_gross, 2),
                    "payment_status": "pagata" if doc_fic.get("is_marked") else "non_pagata",
                    "created_at": now,
                    "updated_at": now,
                }
                # Level 1 already done (fic_scadenze from payments_list)
                # Level 2: Fallback from supplier if no scadenze from FIC
                if not fr_doc["scadenze_pagamento"] and not fr_doc["data_scadenza_pagamento"] and fornitore_id:
                    scadenze_calc = await calc_scadenze_from_supplier(
                        db, fornitore_id, user["user_id"],
                        str(doc_fic.get("date", "")),
                        round(amount_gross, 2)
                    )
                    if scadenze_calc:
                        fr_doc["scadenze_pagamento"] = _enrich_scadenze(scadenze_calc, "fornitore")
                        fr_doc["data_scadenza_pagamento"] = scadenze_calc[-1]["data_scadenza"]
                # Level 3: Default 30 days
                if not fr_doc["scadenze_pagamento"] and not fr_doc["data_scadenza_pagamento"]:
                    doc_date_str = str(doc_fic.get("date", ""))
                    if doc_date_str and amount_gross > 0:
                        try:
                            d0 = date.fromisoformat(doc_date_str)
                            default_scad = (d0 + timedelta(days=30)).isoformat()
                            fr_doc["scadenze_pagamento"] = [{
                                "scadenza_id": f"scd_{uuid.uuid4().hex[:8]}",
                                "numero_rata": 1, "totale_rate": 1, "rata": 1,
                                "data_scadenza": default_scad,
                                "importo": round(amount_gross, 2),
                                "importo_residuo": round(amount_gross, 2),
                                "importo_pagato": 0.0, "modalita_pagamento": "",
                                "stato": "aperta", "pagata": False, "origine": "default_30gg",
                            }]
                            fr_doc["data_scadenza_pagamento"] = default_scad
                        except (ValueError, TypeError):
                            pass

                try:
                    await db.fatture_ricevute.insert_one(fr_doc)
                    imported += 1
                except Exception as e:
                    errors.append(f"Errore importazione doc {fic_id}: {str(e)}")

            # Check pagination
            total_pages = resp.get("last_page", resp.get("total_pages", 1))
            if page >= total_pages:
                break
            page += 1

    except Exception as e:
        logger.error(f"FIC sync error: {e}")
        if imported == 0:
            raise HTTPException(502, f"Errore comunicazione FattureInCloud: {str(e)}")

    # PUNTO 3 Ã¢ÂÂ Salva watermark (solo se completato senza eccezioni fatali)
    await db.sync_state.update_one(
        {"user_id": user["user_id"], "type": "fatture_ricevute"},
        {"$set": {
            "last_sync_at": datetime.now(timezone.utc).isoformat(),
            "last_sync_imported": imported,
            "last_sync_skipped": skipped,
        }},
        upsert=True
    )

    return {
        "message": f"Sincronizzazione completata: {imported} importate, {skipped} giÃÂ  presenti",
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
    }


# Ã¢ÂÂÃ¢ÂÂ Recalculate due dates for existing invoices Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.post("/recalc-scadenze")
async def recalc_scadenze(user: dict = Depends(get_current_user)):
    """Recalculate due dates for all received invoices missing scadenze_pagamento, using supplier payment terms.
    Also tries to link unlinked suppliers by P.IVA/CF and bidirectional name matching.
    Verifies existing links and fixes mismatches."""
    uid = user["user_id"]
    
    # Load all suppliers once for matching
    all_suppliers = await db.clients.find(
        {"user_id": uid},
        {"_id": 0, "client_id": 1, "business_name": 1, "partita_iva": 1, "codice_fiscale": 1,
         "supplier_payment_type_id": 1, "payment_type_id": 1}
    ).to_list(500)
    
    supplier_by_id = {s["client_id"]: s for s in all_suppliers}
    
    # Build lookup indexes (strip IT prefix, lowercase)
    def clean_piva(p):
        if not p or p in ("00000000000", "None"):
            return ""
        return re.sub(r'^IT', '', str(p).strip(), flags=re.IGNORECASE)
    
    suffixes_re = re.compile(r'\b(S\.?R\.?L\.?S?\.?|S\.?P\.?A\.?|S\.?N\.?C\.?|S\.?A\.?S\.?|SOC\.?\s*COOP\.?|UNIPERSONALE|DI\s+\S+.*)\b', re.IGNORECASE)
    stop_words = {'DI', 'DEL', 'DELLA', 'DELLO', 'DEI', 'DEGLI', 'DELLE', 'E', 'ED', 'IL', 'LA', 'LO', 'LE', 'I', 'GLI', 'UN', 'UNA', 'UNO', 'PER', 'IN', 'CON', 'SU', 'TRA', 'FRA',
                  'ITALIA', 'ITALIANA', 'ITALIANO', 'EUROPE', 'SERVIZI', 'SERVICE', 'GROUP', 'INTERNATIONAL'}
    
    def get_name_words(name):
        n = suffixes_re.sub('', (name or "").upper().strip()).strip()
        n = re.sub(r'[.,\s]+$', '', n).strip()
        n = re.sub(r'\s{2,}', ' ', n).strip()
        return [w for w in n.split() if w not in stop_words and len(w) >= 3]
    
    def find_best_name_match(fr_nome):
        """Find supplier by name matching (word-based Jaccard)."""
        nome_words = get_name_words(fr_nome)
        if not nome_words:
            return None
        best_match = None
        best_score = 0
        for s in all_suppliers:
            s_words = get_name_words(s.get("business_name", ""))
            if not s_words:
                continue
            common = set(nome_words) & set(s_words)
            if not common:
                continue
            score = len(common)
            total = max(len(nome_words), len(s_words))
            ratio = score / total if total else 0
            if score >= 1 and ratio >= 0.5 and score > best_score:
                best_score = score
                best_match = s["client_id"]
        return best_match
    
    piva_map = {}  # clean_piva -> [client_ids]
    cf_map = {}
    for s in all_suppliers:
        cp = clean_piva(s.get("partita_iva"))
        if cp:
            piva_map.setdefault(cp, []).append(s["client_id"])
        cf = (s.get("codice_fiscale") or "").strip()
        if cf and cf != "None":
            cf_map[cf] = s["client_id"]
    
    def find_best_match(fr_nome, fr_piva, fr_cf):
        """Find best supplier match: P.IVA with name verification > CF > Name."""
        piva = clean_piva(fr_piva)
        cf = (fr_cf or "").strip()
        
        # P.IVA match with name verification
        if piva and piva in piva_map:
            candidates = piva_map[piva]
            if len(candidates) == 1:
                # Verify name is not wildly different
                s = supplier_by_id.get(candidates[0], {})
                fr_words = get_name_words(fr_nome)
                s_words = get_name_words(s.get("business_name", ""))
                common = set(fr_words) & set(s_words)
                if common or not fr_words or not s_words:
                    return candidates[0]
                # Name mismatch Ã¢ÂÂ try name-based instead
            else:
                # Multiple P.IVA matches Ã¢ÂÂ pick the one with best name match
                best = None
                best_score = -1
                for cid in candidates:
                    s = supplier_by_id.get(cid, {})
                    s_words = get_name_words(s.get("business_name", ""))
                    fr_words = get_name_words(fr_nome)
                    common = set(fr_words) & set(s_words)
                    if len(common) > best_score:
                        best_score = len(common)
                        best = cid
                if best:
                    return best
        
        # CF match with name verification
        if cf and cf in cf_map:
            cid = cf_map[cf]
            s = supplier_by_id.get(cid, {})
            fr_words = get_name_words(fr_nome)
            s_words = get_name_words(s.get("business_name", ""))
            common = set(fr_words) & set(s_words)
            if common or not fr_words or not s_words:
                return cid
            # CF match but name mismatch Ã¢ÂÂ fall through to name matching
        
        # Name match
        return find_best_name_match(fr_nome)
    
    # Phase 0: Verify and fix existing links where names don't match
    relinked_count = 0
    linked_frs = db.fatture_ricevute.find(
        {"user_id": uid, "fornitore_id": {"$nin": [None, ""]},
         "$or": [
             {"scadenze_pagamento": {"$in": [[], None]}},
             {"scadenze_pagamento": {"$exists": False}},
         ]},
        {"_id": 0, "fr_id": 1, "fornitore_id": 1, "fornitore_nome": 1, "fornitore_piva": 1, "fornitore_cf": 1}
    )
    async for fr in linked_frs:
        current_id = fr.get("fornitore_id")
        current_supplier = supplier_by_id.get(current_id, {})
        fr_words = get_name_words(fr.get("fornitore_nome", ""))
        s_words = get_name_words(current_supplier.get("business_name", ""))
        
        # Check if current link has no payment type OR name mismatch
        has_pt = current_supplier.get("supplier_payment_type_id") or current_supplier.get("payment_type_id")
        name_matches = bool(set(fr_words) & set(s_words))
        
        if not has_pt or not name_matches:
            better = find_best_match(
                fr.get("fornitore_nome", ""),
                fr.get("fornitore_piva", ""),
                fr.get("fornitore_cf", "")
            )
            if better and better != current_id:
                better_supplier = supplier_by_id.get(better, {})
                better_pt = better_supplier.get("supplier_payment_type_id") or better_supplier.get("payment_type_id")
                if better_pt or name_matches:
                    await db.fatture_ricevute.update_one(
                        {"fr_id": fr["fr_id"]},
                        {"$set": {"fornitore_id": better, "updated_at": datetime.now(timezone.utc)}}
                    )
                    relinked_count += 1
    
    # Phase 1: Link unlinked FRs
    unlinked = db.fatture_ricevute.find(
        {"user_id": uid, "$or": [{"fornitore_id": None}, {"fornitore_id": ""}]},
        {"_id": 0, "fr_id": 1, "fornitore_nome": 1, "fornitore_piva": 1, "fornitore_cf": 1}
    )
    linked_count = 0
    async for fr in unlinked:
        match_id = find_best_match(
            fr.get("fornitore_nome", ""),
            fr.get("fornitore_piva", ""),
            fr.get("fornitore_cf", "")
        )
        if match_id:
            await db.fatture_ricevute.update_one(
                {"fr_id": fr["fr_id"]},
                {"$set": {"fornitore_id": match_id, "updated_at": datetime.now(timezone.utc)}}
            )
            linked_count += 1

    # Phase 2: Recalculate scadenze for all FR with fornitore_id but no scadenze_pagamento
    cursor = db.fatture_ricevute.find(
        {"user_id": uid, "fornitore_id": {"$nin": [None, ""]},
         "$or": [
             {"scadenze_pagamento": {"$in": [[], None]}},
             {"scadenze_pagamento": {"$exists": False}},
             {"scadenze_pagamento": {"$size": 0}},
         ]},
        {"_id": 0, "fr_id": 1, "fornitore_id": 1, "data_documento": 1, "totale_documento": 1}
    )
    updated = 0
    async for fr in cursor:
        fid = fr.get("fornitore_id")
        if not fid:
            continue
        scadenze_calc = await calc_scadenze_from_supplier(
            db, fid, uid, fr.get("data_documento", ""), fr.get("totale_documento", 0)
        )
        if scadenze_calc:
            await db.fatture_ricevute.update_one(
                {"fr_id": fr["fr_id"]},
                {"$set": {
                    "scadenze_pagamento": scadenze_calc,
                    "data_scadenza_pagamento": scadenze_calc[-1]["data_scadenza"],
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            updated += 1
    return {
        "message": f"Ricollegati {relinked_count}, collegati {linked_count} fornitori, ricalcolate {updated} scadenze",
        "relinked": relinked_count, "linked": linked_count, "updated": updated
    }


# Ã¢ÂÂÃ¢ÂÂ Recalculate single invoice scadenze Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.post("/{fr_id}/recalc-scadenze")
async def recalc_single_scadenze(fr_id: str, user: dict = Depends(get_current_user)):
    """Recalculate payment schedule for a single invoice from supplier payment terms."""
    fr = await db.fatture_ricevute.find_one(
        {"fr_id": fr_id, "user_id": user["user_id"]},
        {"_id": 0, "fornitore_id": 1, "data_documento": 1, "totale_documento": 1}
    )
    if not fr:
        raise HTTPException(404, "Fattura non trovata")
    if not fr.get("fornitore_id"):
        raise HTTPException(400, "Fornitore non collegato. Associa prima il fornitore.")

    scadenze_calc = await calc_scadenze_from_supplier(
        db, fr["fornitore_id"], user["user_id"],
        fr.get("data_documento", ""), fr.get("totale_documento", 0)
    )
    if not scadenze_calc:
        raise HTTPException(400, "Il fornitore non ha condizioni di pagamento configurate nell'anagrafica.")

    await db.fatture_ricevute.update_one(
        {"fr_id": fr_id},
        {"$set": {
            "scadenze_pagamento": scadenze_calc,
            "data_scadenza_pagamento": scadenze_calc[-1]["data_scadenza"],
            "updated_at": datetime.now(timezone.utc),
        }}
    )
    return {"scadenze_pagamento": scadenze_calc, "data_scadenza_pagamento": scadenze_calc[-1]["data_scadenza"]}


# Ã¢ÂÂÃ¢ÂÂ Update scadenze manually Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

class ScadenzaUpdate(BaseModel):
    rata: int
    data_scadenza: str
    importo: float
    pagata: bool = False

@router.put("/{fr_id}/scadenze-pagamento")
async def update_scadenze_pagamento(fr_id: str, scadenze: List[ScadenzaUpdate], user: dict = Depends(get_current_user)):
    """Manually update the payment schedule for a received invoice."""
    fr = await db.fatture_ricevute.find_one({"fr_id": fr_id, "user_id": user["user_id"]}, {"_id": 0, "fr_id": 1})
    if not fr:
        raise HTTPException(404, "Fattura non trovata")

    scadenze_list = [s.model_dump() for s in scadenze]
    last_date = scadenze_list[-1]["data_scadenza"] if scadenze_list else ""

    await db.fatture_ricevute.update_one(
        {"fr_id": fr_id},
        {"$set": {
            "scadenze_pagamento": scadenze_list,
            "data_scadenza_pagamento": last_date,
            "updated_at": datetime.now(timezone.utc),
        }}
    )
    return {"scadenze_pagamento": scadenze_list}


# Ã¢ÂÂÃ¢ÂÂ Scadenziario Dashboard Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.get("/scadenziario/dashboard")
async def get_scadenziario_dashboard(
    user: dict = Depends(get_current_user)
):
    """Aggregated deadline dashboard: payments, documents, commesse milestones."""
    uid = user["user_id"]
    today = date.today().isoformat()
    fine_mese = date(date.today().year, date.today().month + 1 if date.today().month < 12 else 1, 1).isoformat() if date.today().month < 12 else f"{date.today().year + 1}-01-01"

    scadenze = []

    # 1. Payment deadlines from fatture ricevute
    fr_cursor = db.fatture_ricevute.find(
        {"user_id": uid, "payment_status": {"$ne": "pagata"}},
        {"_id": 0, "xml_raw": 0}
    ).sort("data_scadenza_pagamento", 1)
    async for fr in fr_cursor:
        # If structured scadenze_pagamento exists, use individual installments
        fr_scadenze = fr.get("scadenze_pagamento") or []
        if fr_scadenze:
            for s_rata in fr_scadenze:
                if s_rata.get("pagata"):
                    continue
                scad = s_rata.get("data_scadenza", "")
                scadenze.append({
                    "tipo": "pagamento",
                    "id": fr.get("fr_id"),
                    "titolo": f"Fatt. {fr.get('numero_documento', '?')} (Rata {s_rata.get('rata', '?')})",
                    "sottotitolo": fr.get("fornitore_nome", ""),
                    "data_scadenza": scad,
                    "data_documento": fr.get("data_documento", ""),
                    "pagamento": fr.get("condizioni_pagamento", ""),
                    "importo": s_rata.get("importo", 0),
                    "stato": "scaduto" if scad and scad < today else ("in_scadenza" if scad and scad <= fine_mese else "ok"),
                    "link": "/fatture-ricevute",
                    "processata": bool(fr.get("imputazione")),
                })
        else:
            # Legacy: single data_scadenza_pagamento
            scad = fr.get("data_scadenza_pagamento", "")
            scadenze.append({
                "tipo": "pagamento",
                "id": fr.get("fr_id"),
                "titolo": f"Fatt. {fr.get('numero_documento', '?')}",
                "sottotitolo": fr.get("fornitore_nome", ""),
                "data_scadenza": scad,
                "data_documento": fr.get("data_documento", ""),
                "pagamento": fr.get("condizioni_pagamento", ""),
                "importo": fr.get("residuo", fr.get("totale_documento", 0)),
                "stato": "scaduto" if scad and scad < today else ("in_scadenza" if scad and scad <= fine_mese else "ok"),
                "link": "/fatture-ricevute",
                "processata": bool(fr.get("imputazione")),
            })

    # 2. Welder certificate expiries
    async for w in db.welders.find({"is_active": True}, {"_id": 0}):
        for q in w.get("qualifications", []):
            exp = q.get("expiry_date", "")
            if exp:
                scadenze.append({
                    "tipo": "patentino",
                    "id": w.get("welder_id"),
                    "titolo": f"Patentino {q.get('standard', '')}",
                    "sottotitolo": w.get("name", ""),
                    "data_scadenza": exp,
                    "importo": None,
                    "stato": "scaduto" if exp < today else ("in_scadenza" if exp <= fine_mese else "ok"),
                    "link": "/quality-hub/welders",
                })

    # 3. Instrument calibration expiries
    async for inst in db.instruments.find({"status": {"$nin": ["fuori_uso"]}}, {"_id": 0}):
        exp = inst.get("next_calibration", "")
        if exp:
            scadenze.append({
                "tipo": "taratura",
                "id": inst.get("instrument_id"),
                "titolo": f"Taratura {inst.get('name', '')}",
                "sottotitolo": inst.get("serial_number", ""),
                "data_scadenza": exp,
                "importo": None,
                "stato": "scaduto" if exp < today else ("in_scadenza" if exp <= fine_mese else "ok"),
                "link": "/quality-hub/equipment",
            })

    # 4. Commesse delivery deadlines
    async for c in db.commesse.find(
        {"user_id": uid, "stato": {"$nin": ["bozza", "chiuso", "fatturato"]}},
        {"_id": 0, "commessa_id": 1, "numero": 1, "title": 1, "data_consegna": 1}
    ):
        dc = c.get("data_consegna", "")
        if dc:
            scadenze.append({
                "tipo": "consegna",
                "id": c.get("commessa_id"),
                "titolo": f"Consegna {c.get('numero', '')}",
                "sottotitolo": c.get("title", ""),
                "data_scadenza": dc,
                "importo": None,
                "stato": "scaduto" if dc < today else ("in_scadenza" if dc <= fine_mese else "ok"),
                "link": f"/commesse/{c.get('commessa_id')}",
            })

    # 5. Incassi attesi Ã¢ÂÂ scadenze da fatture attive emesse
    incassi_scaduti = 0
    incassi_mese = 0
    seen_invoice_ids = set()

    # 5a. Invoices WITH scadenze_pagamento (structured installments)
    async for inv in db.invoices.find(
        {"user_id": uid, "status": {"$in": ["emessa", "inviata_sdi", "accettata"]}, "scadenze_pagamento": {"$exists": True, "$ne": []}},
        {"_id": 0, "invoice_id": 1, "document_number": 1, "client_id": 1, "scadenze_pagamento": 1, "issue_date": 1, "payment_terms": 1}
    ):
        seen_invoice_ids.add(inv.get("invoice_id"))
        client = await db.clients.find_one({"client_id": inv.get("client_id")}, {"_id": 0, "business_name": 1})
        client_name = client.get("business_name", "") if client else ""
        for s in inv.get("scadenze_pagamento", []):
            if s.get("pagata"):
                continue
            sc_date = s.get("data_scadenza", "")
            stato = "scaduto" if sc_date and sc_date < today else ("in_scadenza" if sc_date and sc_date <= fine_mese else "ok")
            importo = s.get("importo", 0)
            if stato == "scaduto":
                incassi_scaduti += importo
            elif stato == "in_scadenza":
                incassi_mese += importo
            scadenze.append({
                "tipo": "incasso",
                "id": inv.get("invoice_id"),
                "titolo": f"Incasso {inv.get('document_number', '?')} (Rata {s.get('rata', '?')})",
                "sottotitolo": client_name,
                "data_scadenza": sc_date,
                "data_documento": inv.get("issue_date", ""),
                "pagamento": inv.get("payment_terms", ""),
                "importo": importo,
                "stato": stato,
                "link": "/fatturazione",
                "rata": s.get("rata"),
            })

    # 5b. Invoices WITHOUT scadenze_pagamento Ã¢ÂÂ use due_date as single payment
    async for inv in db.invoices.find(
        {"user_id": uid,
         "status": {"$in": ["emessa", "inviata_sdi", "accettata"]},
         "payment_status": {"$ne": "pagata"},
         "$or": [
             {"scadenze_pagamento": {"$exists": False}},
             {"scadenze_pagamento": []},
             {"scadenze_pagamento": None},
         ]},
        {"_id": 0, "invoice_id": 1, "document_number": 1, "client_id": 1,
         "totals": 1, "due_date": 1, "issue_date": 1, "payment_terms": 1}
    ):
        if inv.get("invoice_id") in seen_invoice_ids:
            continue
        sc_date = inv.get("due_date") or inv.get("issue_date") or ""
        importo = inv.get("totals", {}).get("total_document", 0) or 0
        if not importo:
            continue
        stato = "scaduto" if sc_date and sc_date < today else ("in_scadenza" if sc_date and sc_date <= fine_mese else "ok")
        if stato == "scaduto":
            incassi_scaduti += importo
        elif stato == "in_scadenza":
            incassi_mese += importo
        client = await db.clients.find_one({"client_id": inv.get("client_id")}, {"_id": 0, "business_name": 1})
        client_name = client.get("business_name", "") if client else ""
        scadenze.append({
            "tipo": "incasso",
            "id": inv.get("invoice_id"),
            "titolo": f"Incasso {inv.get('document_number', '?')} (Rata 1)",
            "sottotitolo": client_name,
            "data_scadenza": sc_date,
            "data_documento": inv.get("issue_date", ""),
            "pagamento": inv.get("payment_terms", ""),
            "importo": importo,
            "stato": stato,
            "link": "/fatturazione",
            "rata": 1,
        })

    # Sort by date
    scadenze.sort(key=lambda x: x.get("data_scadenza") or "9999-12-31")

    # KPIs
    scadute = [s for s in scadenze if s["stato"] == "scaduto"]
    in_scadenza = [s for s in scadenze if s["stato"] == "in_scadenza"]
    pagamenti_scaduti = sum(s.get("importo", 0) or 0 for s in scadute if s["tipo"] == "pagamento")
    pagamenti_mese = sum(s.get("importo", 0) or 0 for s in in_scadenza if s["tipo"] == "pagamento")

    # Aging buckets (solo scadenze finanziarie: pagamento + incasso)
    today_date = date.today()
    aging = {"0_30": 0, "31_60": 0, "61_90": 0, "over_90": 0}
    aging_incassi = {"0_30": 0, "31_60": 0, "61_90": 0, "over_90": 0}
    for s in scadenze:
        if s.get("stato") != "scaduto" or not s.get("importo"):
            continue
        try:
            scad_date = date.fromisoformat(s["data_scadenza"])
            days = (today_date - scad_date).days
        except (ValueError, TypeError):
            continue
        bucket = "0_30" if days <= 30 else ("31_60" if days <= 60 else ("61_90" if days <= 90 else "over_90"))
        target = aging_incassi if s["tipo"] == "incasso" else aging
        target[bucket] = round(target[bucket] + s["importo"], 2)

    # Fatture da processare (inbox)
    inbox_count = await db.fatture_ricevute.count_documents(
        {"user_id": uid, "imputazione": {"$exists": False}, "status": {"$ne": "pagata"}}
    )

    # Totale acquisti anno
    year = date.today().year
    pipeline = [
        {"$match": {"user_id": uid, "data_documento": {"$regex": f"^{year}"}}},
        {"$group": {"_id": None, "totale": {"$sum": "$totale_documento"}}}
    ]
    agg = await db.fatture_ricevute.aggregate(pipeline).to_list(1)
    totale_anno = agg[0]["totale"] if agg else 0

    return {
        "scadenze": scadenze,
        "kpi": {
            "pagamenti_scaduti": round(pagamenti_scaduti, 2),
            "pagamenti_mese_corrente": round(pagamenti_mese, 2),
            "incassi_scaduti": round(incassi_scaduti, 2),
            "incassi_mese_corrente": round(incassi_mese, 2),
            "totale_acquisti_anno": round(totale_anno, 2),
            "scadenze_totali": len(scadenze),
            "scadute": len(scadute),
            "in_scadenza": len(in_scadenza),
            "inbox_da_processare": inbox_count,
            "aging_pagamenti": aging,
            "aging_incassi": aging_incassi,
        }
    }


# Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ
# EXPORT SCADENZIARIO Ã¢ÂÂ XLSX + PDF
# Ã¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂÃ¢ÂÂ

@router.get("/scadenziario/export/xlsx")
async def export_scadenziario_xlsx(
    tipo: Optional[str] = Query(None, description="pagamento | incasso"),
    stato: Optional[str] = Query(None, description="scaduto | in_scadenza | ok"),
    data_dal: Optional[str] = Query(None),
    data_al: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    """Export scadenziario filtrato in formato Excel."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    # Re-use dashboard data
    dashboard = await get_scadenziario_dashboard(user=user)
    scadenze = dashboard.get("scadenze", [])

    # Apply filters
    items = _filter_scadenze(scadenze, tipo, stato, data_dal, data_al)

    wb = Workbook()
    ws = wb.active
    ws.title = "Scadenziario"

    # Header style
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers = ["#", "Tipo", "Scadenza", "Importo", "Documento", "Data Doc.", "Soggetto", "Stato"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    # Column widths
    widths = [5, 10, 14, 14, 30, 14, 30, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    today_d = date.today()
    scaduto_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ok_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    eur_fmt = '#,##0.00 Ã¢ÂÂ¬'

    for idx, item in enumerate(items, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value="Uscita" if item["tipo"] == "pagamento" else "Entrata").border = border

        scad = item.get("data_scadenza", "")
        ws.cell(row=row, column=3, value=scad).border = border
        c4 = ws.cell(row=row, column=4, value=item.get("importo") or 0)
        c4.number_format = eur_fmt
        c4.border = border
        ws.cell(row=row, column=5, value=item.get("titolo", "")).border = border
        ws.cell(row=row, column=6, value=item.get("data_documento", "")).border = border
        ws.cell(row=row, column=7, value=item.get("sottotitolo", "")).border = border

        # Stato with color
        try:
            scad_date = date.fromisoformat(scad) if scad else None
        except ValueError:
            scad_date = None
        if scad_date:
            days = (today_d - scad_date).days
            if days > 0:
                stato_label = f"Scaduto {days}gg"
                fill = scaduto_fill
            elif days >= -7:
                stato_label = f"In scadenza {abs(days)}gg"
                fill = warn_fill
            else:
                stato_label = f"OK ({abs(days)}gg)"
                fill = ok_fill
        else:
            stato_label = "Ã¢ÂÂ"
            fill = None

        c8 = ws.cell(row=row, column=8, value=stato_label)
        c8.border = border
        if fill:
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = fill

    # Totale row
    tot_row = len(items) + 2
    ws.cell(row=tot_row, column=3, value="TOTALE").font = Font(bold=True)
    c_tot = ws.cell(row=tot_row, column=4, value=sum(i.get("importo", 0) or 0 for i in items))
    c_tot.font = Font(bold=True)
    c_tot.number_format = eur_fmt

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Scadenziario_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/scadenziario/export/pdf")
async def export_scadenziario_pdf(
    tipo: Optional[str] = Query(None),
    stato: Optional[str] = Query(None),
    data_dal: Optional[str] = Query(None),
    data_al: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    """Export scadenziario filtrato in formato PDF (WeasyPrint)."""
    from fastapi.responses import StreamingResponse
    from services.pdf_template import render_pdf
    from io import BytesIO

    dashboard = await get_scadenziario_dashboard(user=user)
    scadenze = dashboard.get("scadenze", [])
    items = _filter_scadenze(scadenze, tipo, stato, data_dal, data_al)

    today_d = date.today()
    totale = sum(i.get("importo", 0) or 0 for i in items)

    def fmt_eur(v):
        if not v: return "Ã¢ÂÂ"
        return f"{v:,.2f} Ã¢ÂÂ¬".replace(",", "X").replace(".", ",").replace("X", ".")

    def fmt_date(d):
        if not d: return ""
        p = d.split("-")
        return f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else d

    rows_html = ""
    for idx, item in enumerate(items, 1):
        scad = item.get("data_scadenza", "")
        try:
            scad_date = date.fromisoformat(scad) if scad else None
        except ValueError:
            scad_date = None

        if scad_date:
            days = (today_d - scad_date).days
            if days > 0:
                bg = "#FEE2E2"
                stato_label = f"Scaduto {days}gg"
            elif days >= -7:
                bg = "#FEF3C7"
                stato_label = f"{abs(days)}gg"
            else:
                bg = "#ECFDF5"
                stato_label = f"{abs(days)}gg"
        else:
            bg = "#FFFFFF"
            stato_label = "Ã¢ÂÂ"

        tipo_label = "Uscita" if item["tipo"] == "pagamento" else "Entrata"
        rows_html += f"""<tr style="background:{bg}">
            <td style="text-align:center">{idx}</td>
            <td>{tipo_label}</td>
            <td>{fmt_date(scad)}</td>
            <td style="text-align:right;font-family:monospace">{fmt_eur(item.get('importo'))}</td>
            <td>{item.get('titolo', '')}</td>
            <td>{fmt_date(item.get('data_documento', ''))}</td>
            <td>{item.get('sottotitolo', '')}</td>
            <td>{stato_label}</td>
        </tr>"""

    # Fetch company info
    company = await db.company_profiles.find_one({"user_id": user["user_id"]}, {"_id": 0}) or {}

    html = f"""<!DOCTYPE html>
    <html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 landscape; margin: 12mm; }}
        body {{ font-family: 'Helvetica Neue', Arial, sans-serif; font-size: 9px; color: #1E293B; }}
        h1 {{ font-size: 16px; margin-bottom: 4px; }}
        .meta {{ font-size: 8px; color: #64748B; margin-bottom: 10px; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #334155; color: white; font-size: 8px; padding: 4px 6px; text-align: left; }}
        td {{ padding: 3px 6px; border-bottom: 1px solid #E2E8F0; font-size: 8px; }}
        .totale {{ font-weight: bold; font-size: 9px; margin-top: 8px; text-align: right; }}
    </style></head><body>
    <h1>Scadenziario Ã¢ÂÂ {company.get('business_name', 'Azienda')}</h1>
    <div class="meta">Generato il {fmt_date(today_d.isoformat())} | {len(items)} scadenze | Totale: {fmt_eur(totale)}</div>
    <table>
        <tr><th>#</th><th>Tipo</th><th>Scadenza</th><th>Importo</th><th>Documento</th><th>Data Doc.</th><th>Soggetto</th><th>Stato</th></tr>
        {rows_html}
    </table>
    <p class="totale">Totale: {fmt_eur(totale)}</p>
    </body></html>"""

    pdf_bytes = render_pdf(html).getvalue()
    buf = BytesIO(pdf_bytes)
    buf.seek(0)

    filename = f"Scadenziario_{today_d.isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _filter_scadenze(scadenze, tipo, stato, data_dal, data_al):
    """Filter scadenze based on query params."""
    items = []
    for s in scadenze:
        if s.get("tipo") not in ("pagamento", "incasso"):
            continue
        if tipo and s["tipo"] != tipo:
            continue
        if stato and s.get("stato") != stato:
            continue
        scad = s.get("data_scadenza", "")
        if data_dal and scad and scad < data_dal:
            continue
        if data_al and scad and scad > data_al:
            continue
        items.append(s)
    return items
